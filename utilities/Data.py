import openpyxl

from utilities.Base import Base


class Data:

    @classmethod
    def getTestData(self, testcaseName):
        data_dict = {}
        book = openpyxl.load_workbook(Base.ROOT_PATH + "/resources/input_data.xlsx")
        sheet = book.active

        # loop every row
        for i in range(1, sheet.max_row+1):
            if sheet.cell(row=i, column=1).value == testcaseName:
                # loop through each column
                for j in range(2, sheet.max_column+1):
                    data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [data_dict]